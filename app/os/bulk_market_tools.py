"""Bulk product maintenance and cross-market clone staging.

Bulk operations mutate only local product data. Cross-market cloning first creates a
local reviewable Product, bridges it into Seller OS, then creates a normal listing
approval request; it never publishes directly.
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from app.db import Product, get_db
from app.os.bridge import migrate_legacy_to_os
from app.os.models import OSProduct
from app.os.operations import request_listing_publish
from app.platforms.commerce_ops_api import fetch_remote_product


BULK_COLUMNS = [
    "product_id", "sku", "source", "source_id", "name", "supply_price", "sell_price",
    "category", "brand", "origin", "material", "images_json", "detail_images_json",
    "options_json", "detail_html", "status",
]


def _json_list(value: Any) -> str:
    if value in (None, ""):
        return "[]"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    text = str(value).strip()
    parsed = json.loads(text)
    if not isinstance(parsed, (list, dict)):
        raise ValueError("JSON 값은 배열 또는 객체여야 합니다.")
    return json.dumps(parsed, ensure_ascii=False)


def build_bulk_product_template_xlsx(limit: int = 1000) -> bytes:
    """Export current local product master as an editable XLSX template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(BULK_COLUMNS)
    with get_db() as db:
        rows = db.query(Product).order_by(Product.id.asc()).limit(max(1, int(limit))).all()
        for p in rows:
            ws.append([
                p.id, p.sku, p.source, p.source_id, p.name, p.supply_price, p.sell_price,
                p.category, p.brand, p.origin, p.material, p.images, p.detail_images,
                p.options, p.detail_html, p.status,
            ])
    guide = wb.create_sheet("guide")
    guide.append(["사용법"])
    guide.append(["기존 상품 수정: product_id 또는 sku를 유지하고 수정할 필드만 변경합니다."])
    guide.append(["신규 상품: product_id는 비우고 sku/source/source_id/name/supply_price/sell_price를 입력합니다."])
    guide.append(["images_json/detail_images_json/options_json은 JSON 배열 형식이어야 합니다."])
    guide.append(["status 권장값: draft / ready / listed"])
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def apply_bulk_product_xlsx(data: bytes, *, allow_create: bool = True) -> dict[str, Any]:
    """Apply an XLSX product sheet with per-row validation and transaction isolation."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if "products" not in wb.sheetnames:
        raise ValueError("products 시트가 없습니다.")
    ws = wb["products"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "updated": 0, "errors": []}
    headers = [str(x or "").strip() for x in rows[0]]
    missing = [x for x in BULK_COLUMNS if x not in headers]
    if missing:
        raise ValueError("필수 컬럼 누락: " + ", ".join(missing))
    idx = {name: headers.index(name) for name in headers}
    stats: dict[str, Any] = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for excel_row, values in enumerate(rows[1:], start=2):
        raw = {k: values[i] if i < len(values) else None for k, i in idx.items()}
        if not any(v not in (None, "") for v in raw.values()):
            continue
        try:
            product_id = int(raw.get("product_id") or 0)
            sku = str(raw.get("sku") or "").strip()
            name = str(raw.get("name") or "").strip()
            if not sku or not name:
                raise ValueError("sku와 name은 필수입니다.")
            supply_price = float(raw.get("supply_price") or 0)
            sell_price = float(raw.get("sell_price") or 0)
            if sell_price <= 0:
                raise ValueError("sell_price는 0보다 커야 합니다.")
            images = _json_list(raw.get("images_json"))
            detail_images = _json_list(raw.get("detail_images_json"))
            options = _json_list(raw.get("options_json"))

            with get_db() as db:
                p = db.query(Product).filter_by(id=product_id).first() if product_id else None
                if not p:
                    p = db.query(Product).filter_by(sku=sku).first()
                created = p is None
                if created:
                    if not allow_create:
                        stats["skipped"] += 1; continue
                    source = str(raw.get("source") or "manual").strip()
                    source_id = str(raw.get("source_id") or sku).strip()
                    p = Product(sku=sku, source=source, source_id=source_id, name=name,
                                supply_price=supply_price, sell_price=sell_price)
                    db.add(p)
                p.sku = sku
                p.source = str(raw.get("source") or p.source or "manual").strip()
                p.source_id = str(raw.get("source_id") or p.source_id or sku).strip()
                p.name = name
                p.supply_price = supply_price
                p.sell_price = sell_price
                p.category = str(raw.get("category") or "")
                p.brand = str(raw.get("brand") or "")
                p.origin = str(raw.get("origin") or "")
                p.material = str(raw.get("material") or "")
                p.images = images
                p.detail_images = detail_images
                p.options = options
                p.detail_html = str(raw.get("detail_html") or "")
                p.status = str(raw.get("status") or "ready").strip().lower()
                if p.status not in {"draft", "ready", "listed"}:
                    raise ValueError(f"지원하지 않는 status: {p.status}")
                db.commit()
            stats["created" if created else "updated"] += 1
        except Exception as exc:
            stats["errors"].append({"row": excel_row, "sku": str(raw.get("sku") or ""), "error": f"{type(exc).__name__}: {exc}"})

    migrate_legacy_to_os()
    return stats


def _extract_remote_product(platform: str, external_product_id: str, raw: dict[str, Any]) -> dict[str, Any]:
    platform = str(platform).lower()
    if platform == "coupang":
        items = raw.get("items") or []
        first = items[0] if items else {}
        images = [str(x.get("vendorPath") or "") for x in (first.get("images") or []) if x.get("vendorPath")]
        contents = first.get("contents") or []
        detail_html = ""
        for block in contents:
            for detail in block.get("contentDetails") or []:
                if str(detail.get("detailType") or "").upper() == "TEXT":
                    detail_html += str(detail.get("content") or "")
        return {
            "name": str(raw.get("sellerProductName") or raw.get("displayProductName") or first.get("itemName") or ""),
            "sell_price": float(first.get("salePrice") or first.get("originalPrice") or 0),
            "category": str(raw.get("displayCategoryCode") or ""),
            "brand": str(raw.get("brand") or ""),
            "origin": "",
            "images": images,
            "detail_images": [],
            "options": [],
            "detail_html": detail_html,
        }
    if platform == "smartstore":
        origin = raw.get("originProduct") or raw
        imgs = origin.get("images") or {}
        rep = ((imgs.get("representativeImage") or {}).get("url") if isinstance(imgs, dict) else "") or ""
        optional = [str(x.get("url") or "") for x in (imgs.get("optionalImages") or []) if x.get("url")] if isinstance(imgs, dict) else []
        detail_attr = origin.get("detailAttribute") or {}
        search = detail_attr.get("naverShoppingSearchInfo") or {}
        return {
            "name": str(origin.get("name") or ""),
            "sell_price": float(origin.get("salePrice") or 0),
            "category": str(origin.get("leafCategoryId") or ""),
            "brand": str(search.get("brandName") or search.get("manufacturerName") or ""),
            "origin": str(((detail_attr.get("originAreaInfo") or {}).get("content")) or ""),
            "images": ([rep] if rep else []) + optional,
            "detail_images": [],
            "options": [],
            "detail_html": str(origin.get("detailContent") or ""),
        }
    raise ValueError(f"지원하지 않는 원본 판매채널: {platform}")


def stage_marketplace_clone(
    source_platform: str,
    external_product_id: str,
    target_platform: str,
    *,
    sell_price_override: int | None = None,
    actor: str = "seller",
) -> dict[str, Any]:
    """Import a remote listing locally and create a target publish approval request."""
    source_platform = str(source_platform).strip().lower()
    target_platform = str(target_platform).strip().lower()
    if source_platform not in {"coupang", "smartstore"} or target_platform not in {"coupang", "smartstore"}:
        return {"ok": False, "error": "지원 판매채널은 coupang, smartstore 입니다."}
    if source_platform == target_platform:
        return {"ok": False, "error": "원본과 대상 판매채널이 같습니다."}
    raw = fetch_remote_product(source_platform, str(external_product_id))
    normalized = _extract_remote_product(source_platform, str(external_product_id), raw)
    name = normalized["name"].strip()
    sell_price = int(sell_price_override or normalized["sell_price"] or 0)
    if not name or sell_price <= 0:
        return {"ok": False, "error": "원본 상품에서 상품명/판매가를 확정하지 못했습니다."}
    sku = f"CLONE-{source_platform[:2].upper()}-{str(external_product_id)}"
    with get_db() as db:
        p = db.query(Product).filter_by(sku=sku).first()
        if not p:
            p = Product(
                sku=sku,
                source=f"market_clone_{source_platform}",
                source_id=str(external_product_id),
                source_url="",
                name=name,
                supply_price=0,
                sell_price=sell_price,
                category=normalized["category"],
                brand=normalized["brand"],
                origin=normalized["origin"],
                material="",
                images=json.dumps(normalized["images"], ensure_ascii=False),
                detail_images=json.dumps(normalized["detail_images"], ensure_ascii=False),
                options=json.dumps(normalized["options"], ensure_ascii=False),
                detail_html=normalized["detail_html"],
                status="ready",
            )
            db.add(p); db.commit(); db.refresh(p)
        legacy_id = int(p.id)
    migrate_legacy_to_os()
    with get_db() as db:
        osp = db.query(OSProduct).filter_by(sku=sku).first()
        if not osp:
            return {"ok": False, "error": "Seller OS Product 브리지 생성에 실패했습니다.", "legacy_product_id": legacy_id}
        os_product_id = int(osp.id)
    approval = request_listing_publish(os_product_id, target_platform, actor=actor)
    return {
        "ok": bool(approval.get("ok")),
        "legacy_product_id": legacy_id,
        "os_product_id": os_product_id,
        "sku": sku,
        "source_platform": source_platform,
        "target_platform": target_platform,
        "approval": approval,
        "warning": "원본 마켓 정보에는 공급처/공급가가 없으므로 자동발주용 SupplierOffer는 별도로 연결·검증해야 합니다.",
    }
